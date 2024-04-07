import openpyxl
import json
from openpyxl import load_workbook
from common import contants
class Case:

    def __init__(self):
        self.id=None
        self.title=None
        self.data=None
        self.expected = None
        self.result = None
        self.actual = None
class Do_Excel:
     file_name=None

     def __init__(self,file_name):
         try:
             #操作的文件
             self.file_name=file_name
             #实例化一个workbooku对象
             self.workbook=openpyxl.load_workbook(file_name)
         except FileExistsError as e:
             print('{0} not found, please check file path'.format(file_name))
             raise e
     #读取测试数据方法
     def get_cases(self,sheet_name):
         sheet_name=sheet_name
         sheet=self.workbook[sheet_name]#获取sheet_name表格
         max_row=sheet.max_row
         cases=[]
         for r in range(2,max_row+1):
             case=Case()#实例化一个Case对象，用来存放测试数据
             case.id = sheet.cell(row=r,column=1).value#取表格中的第R行，第一列的值
             case.title = sheet.cell(row=r, column=2).value
             case.data = sheet.cell(row=r, column=3).value
             case.expected = sheet.cell(row=r, column=4).value
             cases.append(case)#将case中的放到cases中
             if type(case.data) == str:
                 case.data = json.loads(case.data)
         return cases
     #写入测试结果
     def write_resule(self,sheet_name,row,result,actual=None):
         sheet=self.workbook[sheet_name]
         sheet.cell(row,5).value=result #写入
         sheet.cell(row,6).value=actual
         self.workbook.save(filename=self.file_name)

# if __name__=='__main__':
#
#     do=Do_Excel(contants.case).get_cases('login')
#
#     print(do)
#     # do.write_resule('login',2,'c','pass')


